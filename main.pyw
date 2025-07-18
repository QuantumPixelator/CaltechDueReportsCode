import os
import sys
import pandas as pd
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from queue import Queue
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkcalendar import Calendar
from PIL import Image, ImageTk

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x, y, _, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 25
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1, font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hide_tip(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class CaltechDueReportsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Caltech Due Date Reporting")
        self.root.geometry("420x550")  # width x height
        self.root.iconbitmap(sys.executable)  # Ensure to have an icon file selected
        self.root.resizable(False, False)
        self.cancel_flag = False
        self.queue = Queue()
        self.folder_paths = []  # Store multiple folder paths

        # Styling
        self.style = ttk.Style()
        self.style.theme_use('winnative')  # Use the winnative theme
        self.style.configure("TButton", padding=5, font=("Helvetica", 10))
        self.style.map("TButton",
                       background=[("active", "#FF0000")], foreground=[("active", "#FF0000")])  # Hover color

        self.style.configure("TFrame", background="white")  # White frame background
        self.style.configure("TLabel", font=("Helvetica", 10), background="white")
        self.root.configure(bg="white")  # Set root background to white

        # Configure style for Checkbuttons
        self.style.configure("TCheckbutton", background="white")
        self.style.map(
            "TCheckbutton",
            foreground=[("active", "#FF0000")],  # Red text on hover
            bordercolor=[("active", "#FF0000")],  # Red outline on hover
            borderwidth=[("active", 2)],
            relief=[("active", "solid")],        # Solid border on hover
        )

        # Main frame
        main_frame = ttk.Frame(self.root, padding="10", style="TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(0, weight=1)  # Make column 0 expandable
        main_frame.columnconfigure(1, weight=0)  # Column 1 just fits content

        # Folder selection (labels in a frame for perfect alignment)
        folder_label_frame = ttk.Frame(main_frame, style="TFrame")
        folder_label_frame.grid(row=0, column=0, sticky="ew", pady=5)
        folder_label_frame.columnconfigure(0, weight=1)
        ttk.Label(folder_label_frame, text="Selected Folders:", style="TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(folder_label_frame, text="v1.5", foreground="#A0A0A0", font=("TkDefaultFont", 9), style="TLabel").grid(row=0, column=0, sticky="e")

        # Listbox to display selected folders
        self.folder_listbox = tk.Listbox(main_frame, width=63, height=5, bg="white")
        self.folder_listbox.grid(row=1, column=0, padx=5, pady=5)
        
        # Folder buttons frame
        folder_button_frame = ttk.Frame(main_frame, style="TFrame")
        folder_button_frame.grid(row=2, column=0, columnspan=2, pady=5)
        
        ttk.Button(folder_button_frame, text="Add Folder", 
                  command=self.add_folder).pack(side=tk.LEFT, padx=5)
        ttk.Button(folder_button_frame, text="Remove Selected", 
                  command=self.remove_folder).pack(side=tk.LEFT, padx=5)
        ttk.Button(folder_button_frame, text="Clear All", 
                  command=self.clear_folders).pack(side=tk.LEFT, padx=5)

        # Subfolder checkbox
        self.subfolder_var = tk.BooleanVar()
        ttk.Checkbutton(main_frame, text="Include Subfolders",
                       variable=self.subfolder_var, style="TCheckbutton").grid(row=3, column=0, padx=(37, 5), sticky="w", pady=5)

        # Federal government location checkbox
        self.fedgov_var = tk.BooleanVar()
        fedgov_cb = ttk.Checkbutton(main_frame, text="Fed Gov't Location", variable=self.fedgov_var, style="TCheckbutton")
        fedgov_cb.grid(row=3, column=0, padx=(203, 5), pady=5)
        ToolTip(fedgov_cb, "Check if a federal gov't location.")

        # Empty row for spacing
        ttk.Label(main_frame, text=" ").grid(row=4, column=0)

        # Date range (buttons and labels stacked in columns)
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()

        # Frame for date buttons and labels
        date_frame = ttk.Frame(main_frame)
        date_frame.grid(row=5, column=0, columnspan=2, sticky="ew", padx=5, pady=(5,2))
        date_frame.columnconfigure(0, weight=1)
        date_frame.columnconfigure(1, weight=1)

        # Start Date button and label
        start_date_button = ttk.Button(date_frame, text="Start Date", command=lambda: self.pick_date(self.start_date_var), width=10)
        start_date_button.grid(row=0, column=0, sticky="e", padx=(0, 12))
        self.start_date_label = ttk.Label(date_frame, textvariable=self.start_date_var, width=12, relief="sunken", anchor="center")
        self.start_date_label.grid(row=1, column=0, sticky="e", padx=(0, 12), pady=(2,0))

        # End Date button and label
        end_date_button = ttk.Button(date_frame, text="End Date", command=lambda: self.pick_date(self.end_date_var), width=10)
        end_date_button.grid(row=0, column=1, sticky="w")
        self.end_date_label = ttk.Label(date_frame, textvariable=self.end_date_var, width=12, relief="sunken", anchor="center")
        self.end_date_label.grid(row=1, column=1, sticky="w", pady=(2,0))

        # Empty row for spacing
        ttk.Label(main_frame, text=" ").grid(row=7, column=0)

        # Logo addition
        try:
            logo_path = os.path.join(getattr(sys, '_MEIPASS', os.path.abspath(".")), "logo.png") if hasattr(sys, '_MEIPASS') else "logo.png"
            logo_img = Image.open(logo_path)
            aspect_ratio = logo_img.width / logo_img.height
            max_logo_height = 31  # Adjust as needed
            new_height = min(logo_img.height, max_logo_height)
            new_width = int(new_height * aspect_ratio)
            logo_img = logo_img.resize((new_width, new_height), Image.LANCZOS)
            self.logo_photo_date = ImageTk.PhotoImage(logo_img)  # Different variable name
            logo_label_date = ttk.Label(main_frame, image=self.logo_photo_date)
            logo_label_date.grid(row=8, column=0, columnspan=2, sticky="n", pady=(5, 0))
        except Exception as e:
            print(f"Error loading logo: {e}")

        # Empty row for spacing
        ttk.Label(main_frame, text=" ").grid(row=9, column=0)

        # Progress bar (shifted down)
        self.style.configure("red.Horizontal.TProgressbar", troughcolor="white", bordercolor="white", background="#FF0000", lightcolor="#FF0000", darkcolor="#FF0000")
        self.progress = ttk.Progressbar(main_frame, length=400, mode="determinate", style="red.Horizontal.TProgressbar")
        self.progress.grid(row=11, column=0, columnspan=2, pady=10)

        # Status label (shifted down)
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(main_frame, textvariable=self.status_var).grid(row=12, column=0, columnspan=2, pady=5)

        # Buttons (shifted down)
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=13, column=0, columnspan=2, pady=10)
        self.search_button = ttk.Button(button_frame, text="Search", command=self.start_search)
        self.search_button.pack(side=tk.LEFT, padx=5)
        self.cancel_button = ttk.Button(button_frame, text="Cancel", command=self.cancel_search, state="disabled")
        self.cancel_button.pack(side=tk.LEFT, padx=5)

        # Author label (shifted down)
        ttk.Label(main_frame, text="Quantum Pixelator", foreground="#A0A0A0", 
                 font=("TkDefaultFont", 9)).grid(row=14, column=0, columnspan=2, pady=5)

        # Center the window
        self.root.eval('tk::PlaceWindow . center')

    def add_folder(self):
        folder = filedialog.askdirectory()
        if folder and folder not in self.folder_paths:
            self.folder_paths.append(folder)
            self.folder_listbox.insert(tk.END, folder)

    def remove_folder(self):
        try:
            selected = self.folder_listbox.curselection()
            if selected:
                folder = self.folder_listbox.get(selected[0])
                self.folder_paths.remove(folder)
                self.folder_listbox.delete(selected[0])
        except IndexError:
            messagebox.showwarning("Warning", "Please select a folder to remove.")

    def clear_folders(self):
        self.folder_paths.clear()
        self.folder_listbox.delete(0, tk.END)

    def validate_dates(self, start_date, end_date):
        try:
            start = datetime.strptime(start_date, '%m/%d/%Y')
            end = datetime.strptime(end_date, '%m/%d/%Y')
            if start > end:
                messagebox.showerror("Error", "Start date must be before end date.")
                return False
            return True
        except ValueError:
            messagebox.showerror("Error", "Dates must be in MM/DD/YYYY format.")
            return False

    def setup_logging(self, folder_path):
        log_file = os.path.join(folder_path, "error_log.txt")
        logging.basicConfig(filename=log_file, level=logging.INFO,
                           format='%(asctime)s - %(message)s')
        logging.getLogger().handlers[0].flush()
        logging.info("Logging initialized.")
        return log_file

    def get_cell_value(self, ws, row, col, merged_ranges):
        cell = ws.cell(row=row, column=col)
        value = cell.value
        for merged in merged_ranges:
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                value = ws.cell(row=merged.min_row, column=merged.min_col).value
                break
        return value if value is not None else None

    def search_excel_files(self, folder_paths, include_subfolders, start_date, end_date):
        self.cancel_flag = False
        self.status_var.set("Scanning files...")
        self.search_button.config(state="disabled")
        self.cancel_button.config(state="normal")
        self.progress["value"] = 0

        # Setup logging in the first folder
        log_file = self.setup_logging(folder_paths[0])
        results = []
        excel_files = []
        successful_files = 0
        failed_files = 0

        # Collect Excel files from all selected folders
        for folder_path in folder_paths:
            if include_subfolders:
                for root, _, files in os.walk(folder_path):
                    if self.cancel_flag:
                        logging.info("Search cancelled by user.")
                        logging.getLogger().handlers[0].flush()
                        break
                    excel_files.extend(os.path.join(root, f) for f in files
                                      if f.endswith(('.xlsx', '.xls')) and 'summary' in f.lower())
            else:
                excel_files.extend(os.path.join(folder_path, f) for f in os.listdir(folder_path)
                                  if f.endswith(('.xlsx', '.xls')) and 'summary' in f.lower())

        total_files = len(excel_files)
        if total_files == 0:
            logging.info("No summary files were found.")
            logging.getLogger().handlers[0].flush()
            self.queue.put(("done", [], "No summary files were found.", 0, 0))
            return

        # Process each file
        for i, file_path in enumerate(excel_files):
            if self.cancel_flag:
                logging.info("Search cancelled by user.")
                logging.getLogger().handlers[0].flush()
                break
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                merged_ranges = ws.merged_cells.ranges

                # Use alternate cell locations if Federal Government Location is checked
                if self.fedgov_var.get():
                    company_name = self.get_cell_value(ws, 2, 8, merged_ranges)  # H2
                    id_col = 2   # column B
                    desc_col = 5 # column E
                    due_col = 11 # column K
                else:
                    company_name = self.get_cell_value(ws, 2, 7, merged_ranges)  # G2
                    id_col = 2   # column B
                    desc_col = 4 # column D
                    due_col = 9  # column I

                if company_name is None:
                    logging.info(f"Error processing {os.path.basename(file_path)}: Empty or missing value in Company Name cell.")
                    failed_files += 1
                    logging.getLogger().handlers[0].flush()
                    wb.close()
                    continue

                row_count = 0
                valid_row_found = False
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max(id_col, desc_col, due_col)):
                    if self.cancel_flag:
                        logging.info("Search cancelled by user.")
                        logging.getLogger().handlers[0].flush()
                        break
                    if all(cell.value is None for cell in row):
                        continue
                    row_count += 1

                    due_date = self.get_cell_value(ws, row[due_col-1].row, due_col, merged_ranges)
                    if due_date is None:
                        continue

                    try:
                        due_date = datetime.strptime(str(due_date), '%m/%d/%Y')
                        start = datetime.strptime(start_date, '%m/%d/%Y')
                        end = datetime.strptime(end_date, '%m/%d/%Y')
                        if start <= due_date <= end:
                            client_id = self.get_cell_value(ws, row[id_col-1].row, id_col, merged_ranges)
                            description = self.get_cell_value(ws, row[desc_col-1].row, desc_col, merged_ranges)

                            if None in (client_id, description):
                                logging.info(f"Error processing {os.path.basename(file_path)}, row {row[due_col-1].row}: Missing values in ID or Description.")
                                continue

                            results.append({
                                "Client": company_name,
                                "Description": description,
                                "Client ID": client_id,
                                "Due Date": due_date.strftime('%m/%d/%Y')
                            })
                            valid_row_found = True
                    except (ValueError, TypeError) as e:
                        logging.info(f"Error processing {os.path.basename(file_path)}, row {row[due_col-1].row}: Invalid date format in due date cell ({due_date}). Error: {str(e)}")
                        continue

                if valid_row_found:
                    successful_files += 1
                else:
                    failed_files += 1
                    logging.info(f"Error processing {os.path.basename(file_path)}: No valid rows with dates in range found.")

                self.queue.put(("progress", (i + 1) / total_files * 100))
                logging.getLogger().handlers[0].flush()
                wb.close()
            except Exception as e:
                logging.info(f"Error processing {os.path.basename(file_path)}: {str(e)}")
                failed_files += 1
                logging.getLogger().handlers[0].flush()
                if 'wb' in locals():
                    wb.close()
                continue

        summary = f"Search Complete\nSuccessful Files: {successful_files}\nFailed Files: {failed_files}\nDetails in {log_file}"
        logging.info(summary)
        logging.getLogger().handlers[0].flush()
        self.queue.put(("done", results, summary, successful_files, failed_files))

    def start_search(self):
        folder_paths = self.folder_paths
        start_date = self.start_date_var.get()
        end_date = self.end_date_var.get()

        if not folder_paths:
            messagebox.showerror("Error", "Please select at least one folder.")
            return
        if not start_date or not end_date:
            messagebox.showerror("Error", "Please enter both start and end dates.")
            return
        if not self.validate_dates(start_date, end_date):
            return

        # Start search in a separate thread
        threading.Thread(target=self.search_excel_files, 
                       args=(folder_paths, self.subfolder_var.get(), start_date, end_date),
                       daemon=True).start()
        self.root.after(100, self.check_queue)

    def check_queue(self):
        try:
            while not self.queue.empty():
                msg = self.queue.get()
                if msg[0] == "progress":
                    self.progress["value"] = msg[1]
                    self.status_var.set(f"Processing: {int(msg[1])}%")
                elif msg[0] == "done":
                    results, summary, successful_files, failed_files = msg[1], msg[2], msg[3], msg[4]
                    self.search_button.config(state="normal")
                    self.cancel_button.config(state="disabled")
                    self.progress["value"] = 100
                    self.status_var.set("Processing complete.")
                    if results:
                        df = pd.DataFrame(results)
                        self.show_dataframe_window(df, summary)
                    else:
                        messagebox.showinfo("Results", summary)
                    return
        except Exception as e:
            logging.info(f"Queue processing error: {str(e)}")
            logging.getLogger().handlers[0].flush()
            self.status_var.set("Error occurred. See log for details.")
            self.search_button.config(state="normal")
            self.cancel_button.config(state="disabled")
            self.progress["value"] = 0
        if not self.cancel_flag:
            self.root.after(100, self.check_queue)

    def show_dataframe_window(self, df, summary):
        win = tk.Toplevel(self.root)
        win.title("Results Preview")
        win.geometry("700x400")
        frame = ttk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True)

        tree = ttk.Treeview(frame, columns=list(df.columns), show="headings", selectmode="extended")
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor="center")
        for _, row in df.iterrows():
            tree.insert("", tk.END, values=list(row))
        tree.pack(fill=tk.BOTH, expand=True)

        # Add a scrollbar
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Context menu for row actions
        menu = tk.Menu(win, tearoff=0)
        win.bind("<Button-3>", lambda e: menu.post(e.x_root, e.y_root))

        menu.add_command(label="Copy Row(s)", command=lambda: copy_selected_rows())

        def copy_selected_rows():
            selected = tree.selection()
            if selected:
                rows = []
                for iid in selected:
                    values = tree.item(iid, "values")
                    rows.append('\t'.join(str(v) for v in values))
                self.root.clipboard_clear()
                self.root.clipboard_append('\n'.join(rows))

        # Ask to save
        def ask_save():
            win.destroy()
            if messagebox.askyesno("Save Results", "Would you like to save these results to a new Excel file?"):
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    title="Save Results As"
                )
                if file_path:
                    df.to_excel(file_path, index=False, engine='openpyxl')
                    messagebox.showinfo("Success", f"{summary}\nResults saved to {file_path}")

        ttk.Button(win, text="Close", command=ask_save).pack(pady=8)

    def cancel_search(self):
        self.cancel_flag = True
        self.search_button.config(state="normal")
        self.cancel_button.config(state="disabled")
        self.status_var.set("Search cancelled.")
        self.progress["value"] = 0
        logging.info("Search cancelled by user.")
        logging.getLogger().handlers[0].flush()

    def pick_date(self, date_var):
        top = tk.Toplevel(self.root)
        # Determine which variable is being set
        if date_var is self.start_date_var:
            # First day of current month
            today = datetime.today()
            sel_date = today.replace(day=1)
        elif date_var is self.end_date_var:
            # Last day of current month
            today = datetime.today()
            if today.month == 12:
                sel_date = today.replace(day=31)
            else:
                next_month = today.replace(day=28) + timedelta(days=4)
                sel_date = next_month.replace(day=1) - timedelta(days=1)
        else:
            sel_date = datetime.today()

        cal = Calendar(
            top,
            selectmode='day',
            date_pattern='mm/dd/yyyy',
            year=sel_date.year,
            month=sel_date.month,
            day=sel_date.day
        )
        cal.pack(padx=10, pady=10)
        def set_date():
            date_var.set(cal.get_date())
            top.destroy()
        ttk.Button(top, text="Select", command=set_date).pack(pady=5)

if __name__ == "__main__":
    root = tk.Tk()
    app = CaltechDueReportsApp(root)
    root.mainloop()